from decimal import Decimal, InvalidOperation
from django.shortcuts import render, redirect, get_object_or_404
from .models import Purchase, PurchaseItem, Mill, Product, Payment, SaleItem, Sale, Broker
from django.db import transaction
from django.db.models import Q, Sum, F, FloatField, ExpressionWrapper
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
import json
from django.utils import timezone
from io import BytesIO
from django.conf import settings
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib import colors
import qrcode
from reportlab.lib.utils import ImageReader
from num2words import num2words
from reportlab.pdfbase.pdfmetrics import stringWidth
from .function.add_sale import add_sale
from .function.sale_invoice_pdf import sale_invoice_pdf

def dashboard(request):
    return render(request, 'core/dashboard.html')

def add_purchase(request):
    mills = Mill.objects.all().order_by("mill_name")
    products = Product.objects.filter(is_active=True).order_by("rice_name")

    if request.method == "POST":
        mill_id = request.POST.get("mill")
        invoice_no = request.POST.get("invoice_no")
        purchase_date = request.POST.get("purchase_date")

        # multiple row values
        product_ids = request.POST.getlist("product[]")
        bag_weights = request.POST.getlist("bag_weight[]")
        bag_counts = request.POST.getlist("bag_count[]")
        rates = request.POST.getlist("purchase_price[]")  # ✅ rate per KG

        total_amount = 0

        with transaction.atomic():
            purchase = Purchase.objects.create(
                mill_id=mill_id,
                invoice_no=invoice_no,
                purchase_date=purchase_date,
                total_amount=0
            )

            for i in range(len(product_ids)):
                if not product_ids[i]:
                    continue

                bw = int(bag_weights[i] or 0)        # ✅ bag weight (50 default)
                bc = int(bag_counts[i] or 0)         # ✅ bag count
                rate = float(rates[i] or 0)          # ✅ rate per KG

                # ✅ Rule B: amount = bags * bag_weight * rate_per_kg
                line_total = bc * bw * rate
                total_amount += line_total

                PurchaseItem.objects.create(
                    purchase=purchase,
                    product_id=product_ids[i],
                    bag_weight=bw,
                    bag_count=bc,
                    purchase_price=rate
                )

            purchase.total_amount = total_amount
            purchase.save()

        messages.success(request, "✅ Purchase saved successfully!")
        return redirect("purchase_list")

    return render(request, "core/add_purchase.html", {"mills": mills, "products": products})


def purchase_list(request):
    purchases = (
        Purchase.objects.select_related("mill")
        .annotate(
            total_bags=Sum("purchaseitem__bag_count"),
            total_kg=Sum(
                ExpressionWrapper(
                    F("purchaseitem__bag_count") * F("purchaseitem__bag_weight"),
                    output_field=FloatField()
                )
            )
        )
        .order_by("-id")
    )

    return render(request, "core/purchase_list.html", {"purchases": purchases})

def purchase_detail(request, purchase_id):
    purchase = get_object_or_404(Purchase, id=purchase_id)
    items = PurchaseItem.objects.filter(purchase=purchase).select_related("product")

    # ✅ build rows with calculations
    item_rows = []
    total_amount = 0
    for it in items:
        row_kg = (it.bag_weight or 0) * (it.bag_count or 0)
        amount = row_kg * float(it.purchase_price or 0)
        total_amount += amount

        item_rows.append({
            "rice_name": it.product.rice_name,
            "bag_weight": it.bag_weight,
            "bag_count": it.bag_count,
            "row_kg": row_kg,
            "rate_per_kg": it.purchase_price,
            "amount": amount,
        })

    invoice_payments = Payment.objects.filter(
        related_type="purchase",
        purchase=purchase
    ).order_by("-payment_date", "-id")

    invoice_paid = invoice_payments.aggregate(s=Sum("amount"))["s"] or 0
    invoice_due = float(purchase.total_amount) - float(invoice_paid)

    total_bags = items.aggregate(s=Sum("bag_count"))["s"] or 0
    total_kg = items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

    return render(request, "core/purchase_detail.html", {
        "purchase": purchase,
        "item_rows": item_rows,   # ✅ new
        "total_bags": total_bags,
        "total_kg": total_kg,
        "total_amount": total_amount,  # ✅ new
        "invoice_payments": invoice_payments,
        "invoice_paid": invoice_paid,
        "invoice_due": invoice_due,
    })


def edit_purchase(request, purchase_id):
    purchase = get_object_or_404(Purchase, id=purchase_id)
    mills = Mill.objects.all().order_by("mill_name")
    products = Product.objects.filter(is_active=True).order_by("rice_name")

    items = PurchaseItem.objects.filter(purchase=purchase).select_related("product")

    if request.method == "POST":
        mill_id = request.POST.get("mill")
        invoice_no = request.POST.get("invoice_no")
        purchase_date = request.POST.get("purchase_date")

        product_ids = request.POST.getlist("product[]")
        bag_weights = request.POST.getlist("bag_weight[]")
        bag_counts = request.POST.getlist("bag_count[]")
        rates = request.POST.getlist("purchase_price[]")

        total_amount = 0

        with transaction.atomic():
            # Update purchase header
            purchase.mill_id = mill_id
            purchase.invoice_no = invoice_no
            purchase.purchase_date = purchase_date
            purchase.save()

            # Remove old items
            PurchaseItem.objects.filter(purchase=purchase).delete()

            # Insert new items
            for i in range(len(product_ids)):
                if not product_ids[i]:
                    continue

                bw = int(bag_weights[i] or 0)
                bc = int(bag_counts[i] or 0)
                rate = float(rates[i] or 0)

                line_total = bc * bw * rate
                total_amount += line_total

                PurchaseItem.objects.create(
                    purchase=purchase,
                    product_id=product_ids[i],
                    bag_weight=bw,
                    bag_count=bc,
                    purchase_price=rate
                )

            purchase.total_amount = total_amount
            purchase.save()

        messages.success(request, "✅ Purchase updated successfully!")
        return redirect("purchase_detail", purchase_id=purchase.id)

    return render(request, "core/edit_purchase.html", {
        "purchase": purchase,
        "mills": mills,
        "products": products,
        "items": items,
    })

def delete_purchase(request, purchase_id):
    purchase = get_object_or_404(Purchase.objects.select_related("mill"), id=purchase_id)

    if request.method == "POST":
        purchase.delete()  # ✅ automatically deletes PurchaseItem also
        messages.success(request, "🗑 Purchase invoice deleted successfully!")
        return redirect("purchase_list")

    return render(request, "core/delete_purchase.html", {"purchase": purchase})


def add_mill(request):
    if request.method == "POST":
        Mill.objects.create(
            mill_name=request.POST.get("mill_name"),
            owner_name=request.POST.get("owner_name", ""),
            mobile=request.POST.get("mobile"),
            address=request.POST.get("address", ""),
            gst_number=request.POST.get("gst_number", ""),
            opening_balance=request.POST.get("opening_balance") or 0,
        )
        messages.success(request, "✅ Mill saved successfully!")
        return redirect("mill_list")

    return render(request, "core/add_mill.html")

def mill_list(request):
    q = request.GET.get("q", "").strip()

    mills = Mill.objects.all().order_by("-created_at")

    if q:
        mills = mills.filter(
            Q(mill_name__icontains=q) |
            Q(owner_name__icontains=q) |
            Q(mobile__icontains=q)
        )

    return render(request, "core/mill_list.html", {"mills": mills})


def edit_mill(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    if request.method == "POST":
        mill.mill_name = request.POST.get("mill_name")
        mill.owner_name = request.POST.get("owner_name", "")
        mill.mobile = request.POST.get("mobile")
        mill.address = request.POST.get("address", "")
        mill.gst_number = request.POST.get("gst_number", "")
        mill.opening_balance = request.POST.get("opening_balance") or 0
        mill.save()

        messages.success(request, "Mill updated successfully ✅")
        return redirect("mill_list")

    return render(request, "core/edit_mill.html", {"mill": mill})


def delete_mill(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    if request.method == "POST":
        mill.delete()
        messages.success(request, "Mill deleted successfully 🗑️")
        return redirect("mill_list")

    return render(request, "core/delete_mill.html", {"mill": mill})


def product_list(request):
    q = request.GET.get("q", "").strip()

    products = Product.objects.all().order_by("-id")  # ✅ no created_at in your model

    if q:
        products = products.filter(
            Q(rice_name__icontains=q) |
            Q(hsn_code__icontains=q)
        )

    return render(request, "core/product_list.html", {"products": products})


def add_product(request):
    if request.method == "POST":
        Product.objects.create(
            rice_name=request.POST.get("rice_name"),
            hsn_code=request.POST.get("hsn_code", ""),
            gst_percent=request.POST.get("gst_percent") or 0,
            is_active=True if request.POST.get("is_active") == "on" else False,
        )
        messages.success(request, "✅ Product saved successfully!")
        return redirect("product_list")

    return render(request, "core/add_product.html")


def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == "POST":
        product.rice_name = request.POST.get("rice_name")
        product.hsn_code = request.POST.get("hsn_code", "")
        product.gst_percent = request.POST.get("gst_percent") or 0
        product.is_active = True if request.POST.get("is_active") == "on" else False
        product.save()

        messages.success(request, "✅ Product updated successfully!")
        return redirect("product_list")

    return render(request, "core/edit_product.html", {"product": product})


def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == "POST":
        product.delete()
        messages.success(request, "🗑 Product deleted successfully!")
        return redirect("product_list")

    return render(request, "core/delete_product.html", {"product": product})


def product_report(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    total_purchase_bags = (
        PurchaseItem.objects.filter(product=product)
        .aggregate(s=Sum("bag_count"))["s"] or 0
    )

    total_sale_bags = (
        SaleItem.objects.filter(product=product)
        .aggregate(s=Sum("bag_count"))["s"] or 0
    )

    current_stock_bags = total_purchase_bags - total_sale_bags

    purchase_items = (
        PurchaseItem.objects.select_related("purchase", "purchase__mill")
        .filter(product=product)
        .order_by("-purchase__purchase_date")
    )

    return render(request, "core/product_report.html", {
        "product": product,
        "total_purchase_bags": total_purchase_bags,
        "total_sale_bags": total_sale_bags,
        "current_stock_bags": current_stock_bags,
        "purchase_items": purchase_items,
    })



def mill_report_detail(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    # ✅ all purchases of this mill
    purchases = Purchase.objects.filter(mill=mill).order_by("-purchase_date", "-id")

    # ✅ all payments to this mill (direct + invoice wise)
    payments = Payment.objects.filter(
        related_type="purchase",
        mill=mill
    ).order_by("-payment_date", "-id")

    # ✅ totals
    total_purchase = purchases.aggregate(s=Sum("total_amount"))["s"] or 0
    total_paid = payments.aggregate(s=Sum("amount"))["s"] or 0

    balance = float(mill.opening_balance) + float(total_purchase) - float(total_paid)

    # ✅ grand totals (bags + kg across all purchases)
    all_items = PurchaseItem.objects.filter(purchase__mill=mill)

    grand_total_bags = all_items.aggregate(s=Sum("bag_count"))["s"] or 0
    grand_total_kg = all_items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

    # ✅ build purchase rows (each invoice)
    purchase_rows = []
    for p in purchases:
        items = PurchaseItem.objects.filter(purchase=p)

        total_bags = items.aggregate(s=Sum("bag_count"))["s"] or 0
        total_kg = items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

        # invoice-wise paid
        paid = Payment.objects.filter(
            related_type="purchase",
            purchase=p
        ).aggregate(s=Sum("amount"))["s"] or 0

        due = float(p.total_amount) - float(paid)
        if due < 0:
            due = 0

        avg_rate = 0
        if total_kg:
            avg_rate = float(p.total_amount) / float(total_kg)

        purchase_rows.append({
            "id": p.id,
            "purchase_date": p.purchase_date,
            "invoice_no": p.invoice_no,

            "total_bags": total_bags,
            "total_kg": total_kg,
            "avg_rate": round(avg_rate, 2),

            "total_amount": p.total_amount,
            "paid": paid,
            "due": round(due, 2),
        })

    return render(request, "core/mill_report_detail.html", {
        "mill": mill,

        "purchases": purchases,
        "payments": payments,

        "total_purchase": total_purchase,
        "total_paid": total_paid,
        "balance": round(balance, 2),

        "grand_total_bags": grand_total_bags,
        "grand_total_kg": grand_total_kg,

        "purchase_rows": purchase_rows,
    })


def add_mill_payment(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    if request.method == "POST":
        amount = request.POST.get("amount")
        payment_mode = request.POST.get("payment_mode")
        payment_date = request.POST.get("payment_date")
        notes = request.POST.get("notes", "")

        Payment.objects.create(
            related_type="purchase",
            mill=mill,          # ✅ direct mill payment
            purchase=None,      # ✅ not invoice linked
            amount=amount,
            payment_mode=payment_mode,
            payment_date=payment_date,
            notes=notes
        )

        messages.success(request, "✅ Payment saved successfully!")
        return redirect("mill_report_detail", mill_id=mill.id)

    return render(request, "core/add_mill_payment.html", {"mill": mill})


def add_purchase_payment(request, purchase_id):
    purchase = get_object_or_404(Purchase, id=purchase_id)
    mill = purchase.mill

    if request.method == "POST":
        amount = request.POST.get("amount")
        payment_mode = request.POST.get("payment_mode")
        payment_date = request.POST.get("payment_date")
        notes = request.POST.get("notes", "")

        Payment.objects.create(
            related_type="purchase",
            mill=mill,
            purchase=purchase,   # ✅ invoice linked
            amount=amount,
            payment_mode=payment_mode,
            payment_date=payment_date,
            notes=notes
        )

        messages.success(request, "✅ Purchase invoice payment saved!")
        return redirect("purchase_detail", purchase_id=purchase.id)

    return render(request, "core/add_purchase_payment.html", {
        "purchase": purchase,
        "mill": mill
    })

def mill_report_excel(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    # reuse same data from your mill_report_detail logic
    purchases = Purchase.objects.filter(mill=mill).order_by("-purchase_date", "-id")
    payments = Payment.objects.filter(related_type="purchase", mill=mill).order_by("-payment_date", "-id")

    total_purchase = purchases.aggregate(s=Sum("total_amount"))["s"] or 0
    total_paid = payments.aggregate(s=Sum("amount"))["s"] or 0
    balance = float(mill.opening_balance) + float(total_purchase) - float(total_paid)

    all_items = PurchaseItem.objects.filter(purchase__mill=mill)
    grand_total_bags = all_items.aggregate(s=Sum("bag_count"))["s"] or 0
    grand_total_kg = all_items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

    # Purchase rows (invoice wise)
    purchase_rows = []
    for p in purchases:
        items = PurchaseItem.objects.filter(purchase=p)
        total_bags = items.aggregate(s=Sum("bag_count"))["s"] or 0
        total_kg = items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

        paid = Payment.objects.filter(related_type="purchase", purchase=p).aggregate(s=Sum("amount"))["s"] or 0
        due = float(p.total_amount) - float(paid)
        if due < 0:
            due = 0

        avg_rate = round(float(p.total_amount) / float(total_kg), 2) if total_kg else 0

        purchase_rows.append([str(p.purchase_date), p.invoice_no, total_bags, total_kg, avg_rate, float(p.total_amount), float(paid), float(due)])

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mill Report"

    # Header
    ws.append(["Mill Report"])
    ws.append([f"Mill: {mill.mill_name}"])
    ws.append([f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"])
    ws.append([])

    # Summary
    ws.append(["Opening Balance", float(mill.opening_balance)])
    ws.append(["Total Purchase", float(total_purchase)])
    ws.append(["Total Paid", float(total_paid)])
    ws.append(["Balance Due", float(balance)])
    ws.append(["Total Bags Purchased", grand_total_bags])
    ws.append(["Total KG Purchased", float(grand_total_kg)])
    ws.append([])

    # Purchases table
    ws.append(["Date", "Invoice", "Bags", "KG", "Avg Rate/KG", "Amount", "Paid", "Due"])
    for row in purchase_rows:
        ws.append(row)

    # Adjust column width
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Response
    # filename = f"Mill_Report_{mill.mill_name.replace(' ', '_')}.xlsx"
    dt = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{mill.mill_name.strip().replace(' ', '_')}_{dt}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def mill_report_pdf(request, mill_id):
    mill = get_object_or_404(Mill, id=mill_id)

    purchases = Purchase.objects.filter(mill=mill).order_by("purchase_date", "id")
    payments = Payment.objects.filter(
        related_type="purchase",
        mill=mill
    ).order_by("payment_date", "id")

    # totals
    total_purchase = purchases.aggregate(s=Sum("total_amount"))["s"] or 0
    total_paid = payments.aggregate(s=Sum("amount"))["s"] or 0
    balance = float(mill.opening_balance) + float(total_purchase) - float(total_paid)

    all_items = PurchaseItem.objects.filter(purchase__mill=mill)
    grand_total_bags = all_items.aggregate(s=Sum("bag_count"))["s"] or 0
    grand_total_kg = all_items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

    # response
    dt = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{mill.mill_name.strip().replace(' ', '_')}_{dt}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    elements = []

    # ---------- TITLE ----------
    elements.append(Paragraph("<b>Maa Bagwati Bhandar</b>", styles["Title"]))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"<b>Mill:</b> {mill.mill_name}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    # ---------- SUMMARY TABLE ----------
    summary_data = [
        ["Opening Balance", f"Rs {mill.opening_balance}"],
        ["Total Purchase", f"Rs {total_purchase}"],
        ["Total Paid", f"Rs {total_paid}"],
        ["Balance Due", f"Rs {round(balance, 2)}"],
        ["Total Bags Purchased", str(grand_total_bags)],
        ["Total KG Purchased", str(grand_total_kg)],
    ]

    summary_table = Table(summary_data, colWidths=[220, 250])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))

    elements.append(Paragraph("<b>Summary</b>", styles["Heading2"]))
    elements.append(summary_table)
    elements.append(Spacer(1, 14))

    # ---------- PURCHASES TABLE ----------
    elements.append(Paragraph("<b>Purchases (Invoice-wise)</b>", styles["Heading2"]))

    purchase_data = [[
        "Date", "Invoice", "Bags", "KG", "Rate/KG", "Amount", "Paid", "Due"
    ]]

    for p in purchases:
        items = PurchaseItem.objects.filter(purchase=p)
        total_bags = items.aggregate(s=Sum("bag_count"))["s"] or 0
        total_kg = items.aggregate(s=Sum(F("bag_weight") * F("bag_count")))["s"] or 0

        paid = Payment.objects.filter(related_type="purchase", purchase=p).aggregate(s=Sum("amount"))["s"] or 0
        due = float(p.total_amount) - float(paid)
        if due < 0:
            due = 0

        rate = round(float(p.total_amount) / float(total_kg), 2) if total_kg else 0

        purchase_data.append([
            str(p.purchase_date),
            p.invoice_no,
            str(total_bags),
            str(total_kg),
            f"Rs {rate}",
            f"Rs {p.total_amount}",
            f"Rs {paid}",
            f"Rs {round(due, 2)}"
        ])

        # ✅ Multiple payments under same invoice
        inv_pays = Payment.objects.filter(related_type="purchase", purchase=p).order_by("payment_date", "id")

        if inv_pays.exists():
            purchase_data.append(["", "", "", "", "", "", "", ""])  # empty row
            purchase_data.append(["", "Payments for Invoice:", "", "", "", "", "", ""])

            for pay in inv_pays:
                purchase_data.append([
                    "",
                    f"- {pay.payment_date} ({pay.payment_mode})",
                    "",
                    "",
                    "",
                    "",
                    f"Rs {pay.amount}",
                    pay.notes or ""
                ])

            purchase_data.append(["", "", "", "", "", "", "", ""])  # separator row

    purchase_table = Table(purchase_data, repeatRows=1, colWidths=[65, 85, 45, 45, 60, 70, 60, 70])
    purchase_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elements.append(purchase_table)
    elements.append(Spacer(1, 14))

    # ---------- PAYMENTS TABLE ----------
    elements.append(PageBreak())
    elements.append(Paragraph("<b>All Payments (Direct + Invoice-wise)</b>", styles["Heading2"]))

    pay_data = [["Date", "Mode", "Invoice", "Amount"]]

    for pay in payments:
        invoice_txt = "-"
        if pay.purchase:
            invoice_txt = pay.purchase.invoice_no

        pay_data.append([
            str(pay.payment_date),
            pay.payment_mode,
            invoice_txt,
            f"Rs {pay.amount}"
        ])

    pay_table = Table(pay_data, repeatRows=1, colWidths=[70, 70, 90, 70, 220])
    pay_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (3, 1), (3, -1), "RIGHT"),
    ]))

    elements.append(pay_table)

    # build PDF
    doc.build(elements)
    return response


def sale_list(request):
    q = request.GET.get("q", "")
    sales = Sale.objects.all().order_by("-sale_date", "-id")

    if q:
        sales = sales.filter(customer_name__icontains=q)

    return render(request, "core/sale_list.html", {"sales": sales, "q": q})

def sale_list(request):
    q = request.GET.get("q", "").strip()

    sales = Sale.objects.select_related("broker").all().order_by("-sale_date", "-id")

    if q:
        sales = sales.filter(
            Q(customer_name__icontains=q) |
            Q(invoice_no__icontains=q) |
            Q(broker__broker_name__icontains=q)
        )

    rows = []
    for s in sales:
        paid = Payment.objects.filter(
            related_type="sale",
            sale=s
        ).aggregate(x=Sum("amount"))["x"] or 0

        due = float(s.total_amount) - float(paid)

        if due <= 0:
            status = "PAID"
        elif paid > 0:
            status = "PARTIAL"
        else:
            status = "DUE"

        rows.append({
            "id": s.id,
            "sale_date": s.sale_date,
            "invoice_no": s.invoice_no,
            "customer_name": s.customer_name,
            "broker": s.broker,
            "total_amount": s.total_amount,
            "paid": paid,
            "due": round(due, 2),
            "status": status,
        })

    return render(request, "core/sale_list.html", {
        "rows": rows,
        "q": q,
    })


from django.db.models import Sum

def sale_detail(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)

    # internal breakup items (BUY cost)
    items = SaleItem.objects.filter(sale=sale).select_related("mill", "product")

    # Rice payments only
    payments = Payment.objects.filter(
        related_type="sale",
        sale=sale
    ).order_by("-payment_date", "-id")

    # ✅ selling side totals (rice invoice)
    rice_total = float(sale.taxable_amount) + float(sale.gst_amount)

    # total kg from sale header
    total_kg = float(sale.total_quantity_kg or 0)

    # ✅ selling rate per kg (auto)
    selling_rate_per_kg = 0
    if total_kg > 0:
        selling_rate_per_kg = float(sale.taxable_amount) / total_kg

    # ✅ total bags from breakup rows
    total_bags = items.aggregate(s=Sum("bag_count"))["s"] or 0

    # ✅ rice received = advance + other payments
    paid_extra = payments.aggregate(s=Sum("amount"))["s"] or 0
    rice_received_total = float(sale.advance_received) + float(paid_extra)
    rice_due = rice_total - rice_received_total
    if rice_due < 0:
        rice_due = 0

    # ✅ transport due (separate)
    transport_due = float(sale.transport_charge) - (
        float(sale.transport_paid_by_dealer) + float(sale.transport_paid_by_customer)
    )
    if transport_due < 0:
        transport_due = 0

    # ✅ buy cost total from breakup rows
    buy_cost_total = items.aggregate(s=Sum("amount"))["s"] or 0

    # ✅ profit estimate (rice selling - buy cost)
    profit_estimate = rice_total - float(buy_cost_total)

    return render(request, "core/sale_detail.html", {
        "sale": sale,
        "items": items,

        # Rice invoice (selling)
        "total_bags": total_bags,
        "total_kg": round(total_kg, 2),
        "selling_rate_per_kg": round(selling_rate_per_kg, 2),
        "rice_total": round(rice_total, 2),
        "paid_extra": round(float(paid_extra), 2),
        "rice_received_total": round(rice_received_total, 2),
        "rice_due": round(rice_due, 2),

        # Transport
        "transport_due": round(transport_due, 2),

        # Internal (buying)
        "buy_cost_total": round(float(buy_cost_total), 2),
        "profit_estimate": round(float(profit_estimate), 2),

        # payments table
        "payments": payments,
    })

def add_sale_payment(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)

    if request.method == "POST":
        amount = request.POST.get("amount")
        payment_mode = request.POST.get("payment_mode")
        payment_date = request.POST.get("payment_date")
        notes = request.POST.get("notes", "")

        Payment.objects.create(
            related_type="sale",
            sale=sale,
            amount=amount,
            payment_mode=payment_mode,
            payment_date=payment_date,
            notes=notes
        )

        messages.success(request, "✅ Sale payment added!")
        return redirect("sale_detail", sale_id=sale.id)

    return render(request, "core/add_sale_payment.html", {"sale": sale})

def generate_sale_invoice_no():
    # Example: SAL-20260118-0001
    today = datetime.now().strftime("%Y%m%d")
    last = Sale.objects.filter(invoice_no__startswith=f"SAL-{today}").order_by("-id").first()
    if last and last.invoice_no:
        try:
            last_seq = int(last.invoice_no.split("-")[-1])
        except:
            last_seq = 0
    else:
        last_seq = 0
    return f"SAL-{today}-{last_seq+1:04d}"


# def add_sale(request):
    products = Product.objects.filter(is_active=True).order_by("rice_name")
    brokers = Broker.objects.all().order_by("broker_name")

    # Purchase stock list for dropdown
    purchase_items = (
        PurchaseItem.objects
        .select_related("purchase", "purchase__mill", "product")
        .order_by("-purchase__purchase_date", "-id")
    )

    purchase_items_json = json.dumps([
        {
            "id": pi.id,
            "product_id": pi.product_id,
            "bag_weight": pi.bag_weight,
            "label": f"{pi.purchase.invoice_no} / {pi.purchase.mill.mill_name} / Buy ₹{pi.purchase_price}/KG / {pi.bag_weight}kg bag",
        }
        for pi in purchase_items
    ])

    # -------------------------
    # GET → show form
    # -------------------------
    if request.method == "GET":
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    # -------------------------
    # POST → review or save
    # -------------------------
    # step = request.POST.get("step", "review")  # "review" or "save"
    step = request.POST.get("step", "review")
    if step != "review":
        step = "review"
    # Common fields
    sale_date = request.POST.get("sale_date")
    customer_name = request.POST.get("customer_name")
    customer_gst = request.POST.get("customer_gst", "")
    broker_id = request.POST.get("broker_id") or None

    vehicle_number = request.POST.get("vehicle_number")
    driver_name = request.POST.get("driver_name")
    transporter_name = request.POST.get("transporter_name")

    # Rice selling fields
    try:
        product_id = int(request.POST.get("product_id") or 0)
    except:
        product_id = 0

    bag_weight = Decimal(request.POST.get("bag_weight") or "0")
    total_bags = int(request.POST.get("total_bags") or 0)

    rate_per_kg = Decimal(request.POST.get("rate_per_kg") or "0")
    gst_percent = Decimal(request.POST.get("gst_percent") or "0")
    advance_received = Decimal(request.POST.get("advance_received") or "0")

    # Transport fields
    transport_rate_per_ton = Decimal(request.POST.get("transport_rate_per_ton") or "0")
    transport_paid_by_dealer = Decimal(request.POST.get("transport_paid_by_dealer") or "0")
    transport_paid_by_customer = Decimal(request.POST.get("transport_paid_by_customer") or "0")

    # Internal breakup fields
    purchase_item_ids = request.POST.getlist("purchase_item[]")
    row_bags_list = request.POST.getlist("row_bags[]")

    # -------------------------
    # Validations
    # -------------------------
    if not sale_date or not customer_name:
        messages.error(request, "Sale Date and Customer Name are required.")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    if product_id <= 0:
        messages.error(request, "Please select a Product.")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    if total_bags <= 0 or bag_weight <= 0:
        messages.error(request, "Total Bags and Bag Weight must be greater than 0.")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    if len(purchase_item_ids) != len(row_bags_list):
        messages.error(request, "Internal breakup rows are invalid.")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    # count only valid rows (pid + bags>0)
    valid_rows = 0
    breakup_sum = 0
    for pid, bags_str in zip(purchase_item_ids, row_bags_list):
        bags = int(bags_str or 0)
        if pid and bags > 0:
            valid_rows += 1
            breakup_sum += bags

    if valid_rows == 0:
        messages.error(request, "Please add at least 1 breakup row (Purchase Stock + Bags).")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    if breakup_sum != total_bags:
        messages.error(request, f"Internal breakup bags ({breakup_sum}) must match Total Bags ({total_bags}).")
        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,
        })

    # -------------------------
    # Calculations
    # -------------------------
    total_kg = Decimal(total_bags) * bag_weight

    taxable_amount = total_kg * rate_per_kg
    gst_amount = (taxable_amount * gst_percent) / Decimal("100")
    rice_total = taxable_amount + gst_amount

    rice_due = rice_total - advance_received
    if rice_due < 0:
        rice_due = Decimal("0")

    total_ton = total_kg / Decimal("1000")
    transport_charge = total_ton * transport_rate_per_ton

    transport_due = transport_charge - (transport_paid_by_dealer + transport_paid_by_customer)
    if transport_due < 0:
        transport_due = Decimal("0")

    grand_total = rice_total + transport_charge

    # -------------------------
    # Build breakup rows (for review + save)
    # -------------------------
    selected_ids = [int(pid) for pid, bags_str in zip(purchase_item_ids, row_bags_list) if pid and int(bags_str or 0) > 0]

    pi_map = {
        pi.id: pi
        for pi in PurchaseItem.objects.select_related("purchase", "purchase__mill", "product").filter(id__in=selected_ids)
    }

    breakup_rows = []
    buy_cost_total = Decimal("0")

    for pid, bags_str in zip(purchase_item_ids, row_bags_list):
        if not pid:
            continue
        bags = int(bags_str or 0)
        if bags <= 0:
            continue

        pi = pi_map.get(int(pid))
        if not pi:
            continue

        if not pi.bag_weight:
            messages.error(request, "Selected purchase stock has no bag weight.")
            return render(request, "core/add_sale.html", {
                "products": products,
                "brokers": brokers,
                "purchase_items_json": purchase_items_json,
            })

        bw = Decimal(str(pi.bag_weight))
        row_kg = bw * Decimal(bags)
        buy_rate = Decimal(str(pi.purchase_price or 0))
        row_amount = row_kg * buy_rate

        buy_cost_total += row_amount

        breakup_rows.append({
            "purchase_item_id": pi.id,
            "invoice_no": pi.purchase.invoice_no,
            "mill_name": pi.purchase.mill.mill_name,
            "mill_id": pi.purchase.mill_id,
            "product_name": pi.product.rice_name,
            "product_id": pi.product_id,
            "bag_weight": int(pi.bag_weight),
            "bags": bags,
            "kg": row_kg,
            "buy_rate": buy_rate,
            "amount": row_amount,
        })

    profit_estimate = rice_total - buy_cost_total

    # -------------------------
    # REVIEW (no save)
    # -------------------------
    if step == "review":
        return render(request, "core/sale_review.html", {
            "sale_date": sale_date,
            "customer_name": customer_name,
            "customer_gst": customer_gst,
            "broker_id": broker_id,

            "vehicle_number": vehicle_number,
            "driver_name": driver_name,
            "transporter_name": transporter_name,

            "product_id": product_id,
            "bag_weight": bag_weight,
            "total_bags": total_bags,
            "rate_per_kg": rate_per_kg,
            "gst_percent": gst_percent,
            "advance_received": advance_received,

            "total_kg": total_kg,
            "taxable_amount": taxable_amount,
            "gst_amount": gst_amount,
            "rice_total": rice_total,
            "rice_due": rice_due,

            "transport_rate_per_ton": transport_rate_per_ton,
            "transport_paid_by_dealer": transport_paid_by_dealer,
            "transport_paid_by_customer": transport_paid_by_customer,
            "transport_charge": transport_charge,
            "transport_due": transport_due,

            "grand_total": grand_total,
            "breakup_rows": breakup_rows,
            "buy_cost_total": buy_cost_total,
            "profit_estimate": profit_estimate,

            # used for hidden inputs on confirm
            "raw_post": request.POST,
        })

    # -------------------------
    # SAVE (confirm)
    # -------------------------
    invoice_no = generate_sale_invoice_no()

    with transaction.atomic():
        sale = Sale.objects.create(
            invoice_no=invoice_no,

            customer_name=customer_name,
            customer_gst=customer_gst,
            broker_id=broker_id,

            sale_date=sale_date,
            vehicle_number=vehicle_number,
            driver_name=driver_name,
            transporter_name=transporter_name,

            transport_charge=transport_charge,
            transport_paid_by_dealer=transport_paid_by_dealer,
            transport_paid_by_customer=transport_paid_by_customer,

            total_quantity_kg=total_kg,
            taxable_amount=taxable_amount,
            gst_percent=gst_percent,
            gst_amount=gst_amount,

            total_amount=grand_total,

            advance_received=advance_received,
            balance_amount=rice_due,
        )

        # ✅ Save breakup as BUY COST rows (NO NULL fields)
        for row in breakup_rows:
            SaleItem.objects.create(
                sale=sale,
                product_id=row["product_id"],
                mill_id=row["mill_id"],
                bag_weight=row["bag_weight"],
                bag_count=row["bags"],
                rate_per_kg=row["buy_rate"],
                total_weight=row["kg"],
                amount=row["amount"],
            )

    messages.success(request, f"✅ Sale saved: {invoice_no}")
    return redirect("sale_detail", sale_id=sale.id)





# def _d(v, default="0"):
#     """Safe Decimal conversion."""
#     try:
#         if v is None or str(v).strip() == "":
#             return Decimal(default)
#         return Decimal(str(v).strip())
#     except (InvalidOperation, ValueError, TypeError):
#         return Decimal(default)


# def add_sale(request):
    products = Product.objects.filter(is_active=True).order_by("rice_name")
    brokers = Broker.objects.all().order_by("broker_name")

    purchase_items = (
        PurchaseItem.objects
        .select_related("purchase", "purchase__mill", "product")
        .order_by("-purchase__purchase_date", "-id")
    )

    purchase_items_json = json.dumps([
        {
            "id": pi.id,
            "product_id": pi.product_id,
            "bag_weight": pi.bag_weight,
            "label": f"{pi.purchase.invoice_no} / {pi.purchase.mill.mill_name} / Buy ₹{pi.purchase_price}/KG / {pi.bag_weight}kg bag",
        }
        for pi in purchase_items
    ])

    # -------------------------
    # GET → show form (prefill from session draft)
    # -------------------------
    if request.method == "GET":
        draft = request.session.get("sale_draft") or {}
        draft_lists = request.session.get("sale_draft_lists") or {}

        return render(request, "core/add_sale.html", {
            "products": products,
            "brokers": brokers,
            "purchase_items_json": purchase_items_json,

            # ✅ Prefill
            "draft": draft,
            "draft_purchase_items": draft_lists.get("purchase_item", []),
            "draft_row_bags": draft_lists.get("row_bags", []),
        })

    # -------------------------
    # POST
    # -------------------------
    step = request.POST.get("step", "").strip()

    # ✅ Step 2: Confirm & Save (from SESSION, not from hidden inputs)
    if step == "save":
        draft = request.session.get("sale_draft") or {}
        draft_lists = request.session.get("sale_draft_lists") or {}

        if not draft:
            messages.error(request, "Draft not found. Please fill the sale form again.")
            return redirect("add_sale")

        # -------- Read scalar fields from draft --------
        sale_date = draft.get("sale_date") or str(timezone.now().date())
        customer_name = (draft.get("customer_name") or "").strip()
        customer_gst = (draft.get("customer_gst") or "").strip()
        broker_id = draft.get("broker_id") or None

        vehicle_number = (draft.get("vehicle_number") or "").strip()
        driver_name = (draft.get("driver_name") or "").strip()
        transporter_name = (draft.get("transporter_name") or "").strip()

        product_id = draft.get("product_id") or None
        bag_weight = int(draft.get("bag_weight") or 0)
        total_bags = int(draft.get("total_bags") or 0)

        rate_per_kg = _d(draft.get("rate_per_kg"), "0")
        gst_percent = int(draft.get("gst_percent") or 0)
        advance_received = _d(draft.get("advance_received"), "0")

        transport_rate_per_ton = _d(draft.get("transport_rate_per_ton"), "0")
        transport_paid_by_dealer = _d(draft.get("transport_paid_by_dealer"), "0")
        transport_paid_by_customer = _d(draft.get("transport_paid_by_customer"), "0")

        # -------- Basic validation --------
        if not customer_name or not product_id or total_bags <= 0 or bag_weight <= 0:
            messages.error(request, "Missing required fields in draft. Please edit and try again.")
            return redirect("add_sale")

        # -------- Compute totals --------
        total_kg = Decimal(total_bags) * Decimal(bag_weight)
        taxable_amount = total_kg * rate_per_kg
        gst_amount = (taxable_amount * Decimal(gst_percent)) / Decimal("100")
        rice_total = taxable_amount + gst_amount

        rice_due = rice_total - advance_received

        # Transport charge based on ton: (kg / 1000) * rate_per_ton
        transport_charge = (total_kg / Decimal("1000")) * transport_rate_per_ton
        transport_due = transport_charge - (transport_paid_by_dealer + transport_paid_by_customer)

        grand_total = rice_total + transport_charge

        # -------- Breakup arrays from session --------
        purchase_item_ids = draft_lists.get("purchase_item", [])
        row_bags_list = draft_lists.get("row_bags", [])

        breakup_rows = []
        buy_cost_total = Decimal("0")

        # Validate arrays length
        if len(purchase_item_ids) != len(row_bags_list):
            messages.error(request, "Breakup rows mismatch. Please edit and try again.")
            return redirect("add_sale")

        # Build breakup rows and compute buy cost
        for pid, bags_str in zip(purchase_item_ids, row_bags_list):
            bags = int(bags_str or 0)
            if bags <= 0:
                continue

            pi = PurchaseItem.objects.select_related("purchase", "purchase__mill").get(id=int(pid))
            kg = Decimal(bags) * Decimal(pi.bag_weight)
            amount = kg * pi.purchase_price

            buy_cost_total += amount
            breakup_rows.append({
                "purchase_item": pi,
                "purchase_item_id": pi.id,
                "invoice_no": pi.purchase.invoice_no,
                "mill_name": pi.purchase.mill.mill_name,
                "bag_weight": pi.bag_weight,
                "bags": bags,
                "kg": kg,
                "buy_rate": pi.purchase_price,
                "amount": amount,
            })

        profit_estimate = rice_total - buy_cost_total

        # -------- Save Sale --------
        broker = Broker.objects.filter(id=broker_id).first() if broker_id else None
        product = Product.objects.get(id=int(product_id))

        sale = Sale.objects.create(
            invoice_no=f"SALE-{timezone.now().strftime('%Y%m%d%H%M%S')}",
            customer_name=customer_name,
            customer_gst=customer_gst,
            sale_date=sale_date,
            vehicle_number=vehicle_number,
            driver_name=driver_name,
            transporter_name=transporter_name,

            total_quantity_kg=total_kg,
            taxable_amount=taxable_amount,
            gst_percent=gst_percent,
            gst_amount=gst_amount,

            # ✅ Keep this as RICE total (recommended)
            total_amount=rice_total,

            advance_received=advance_received,
            balance_amount=rice_due,

            broker=broker,

            # Transport separate
            transport_charge=transport_charge,
            paid_by_dealer=transport_paid_by_dealer,
            paid_by_customer=transport_paid_by_customer,
        )

        # Save SaleItems breakup rows
        # (This assumes your SaleItem has these fields, adjust if different)
        for r in breakup_rows:
            SaleItem.objects.create(
                sale=sale,
                product=product,
                mill=r["purchase_item"].purchase.mill,
                bag_weight=r["bag_weight"],
                bag_count=r["bags"],
                rate_per_kg=rate_per_kg,  # selling rate
                total_weight=r["kg"],
                amount=(r["kg"] * rate_per_kg),
            )

        # ✅ Clear draft after successful save
        request.session.pop("sale_draft", None)
        request.session.pop("sale_draft_lists", None)
        request.session.modified = True

        messages.success(request, "Sale saved successfully ✅")
        return redirect("sale_detail", sale.id)

    # ✅ Step 1: Review (store draft in session + render review)
    # This runs when POST comes from the form and step != save

    # Save draft to session
    request.session["sale_draft"] = dict(request.POST.items())
    request.session["sale_draft_lists"] = {
        "purchase_item": request.POST.getlist("purchase_item[]"),
        "row_bags": request.POST.getlist("row_bags[]"),
    }
    request.session.modified = True

    # Build review page context from POST (same as before)
    sale_date = request.POST.get("sale_date") or str(timezone.now().date())
    customer_name = request.POST.get("customer_name", "")
    customer_gst = request.POST.get("customer_gst", "")
    broker_id = request.POST.get("broker_id") or ""

    vehicle_number = request.POST.get("vehicle_number", "")
    driver_name = request.POST.get("driver_name", "")
    transporter_name = request.POST.get("transporter_name", "")

    product_id = request.POST.get("product_id") or ""
    bag_weight = int(request.POST.get("bag_weight") or 0)
    total_bags = int(request.POST.get("total_bags") or 0)

    rate_per_kg = _d(request.POST.get("rate_per_kg"), "0")
    gst_percent = int(request.POST.get("gst_percent") or 0)
    advance_received = _d(request.POST.get("advance_received"), "0")

    transport_rate_per_ton = _d(request.POST.get("transport_rate_per_ton"), "0")
    transport_paid_by_dealer = _d(request.POST.get("transport_paid_by_dealer"), "0")
    transport_paid_by_customer = _d(request.POST.get("transport_paid_by_customer"), "0")

    total_kg = Decimal(total_bags) * Decimal(bag_weight)
    taxable_amount = total_kg * rate_per_kg
    gst_amount = (taxable_amount * Decimal(gst_percent)) / Decimal("100")
    rice_total = taxable_amount + gst_amount
    rice_due = rice_total - advance_received

    transport_charge = (total_kg / Decimal("1000")) * transport_rate_per_ton
    transport_due = transport_charge - (transport_paid_by_dealer + transport_paid_by_customer)

    grand_total = rice_total + transport_charge

    purchase_item_ids = request.POST.getlist("purchase_item[]")
    row_bags_list = request.POST.getlist("row_bags[]")

    breakup_rows = []
    buy_cost_total = Decimal("0")

    if len(purchase_item_ids) == len(row_bags_list):
        for pid, bags_str in zip(purchase_item_ids, row_bags_list):
            bags = int(bags_str or 0)
            if bags <= 0:
                continue
            pi = PurchaseItem.objects.select_related("purchase", "purchase__mill").get(id=int(pid))
            kg = Decimal(bags) * Decimal(pi.bag_weight)
            amount = kg * pi.purchase_price
            buy_cost_total += amount
            breakup_rows.append({
                "purchase_item_id": pi.id,
                "invoice_no": pi.purchase.invoice_no,
                "mill_name": pi.purchase.mill.mill_name,
                "bag_weight": pi.bag_weight,
                "bags": bags,
                "kg": kg,
                "buy_rate": pi.purchase_price,
                "amount": amount,
            })

    profit_estimate = rice_total - buy_cost_total

    return render(request, "core/sale_review.html", {
        "sale_date": sale_date,
        "customer_name": customer_name,
        "customer_gst": customer_gst,
        "broker_id": broker_id,

        "vehicle_number": vehicle_number,
        "driver_name": driver_name,
        "transporter_name": transporter_name,

        "product_id": product_id,
        "bag_weight": bag_weight,
        "total_bags": total_bags,
        "total_kg": total_kg,
        "rate_per_kg": rate_per_kg,

        "gst_percent": gst_percent,
        "taxable_amount": taxable_amount,
        "gst_amount": gst_amount,
        "rice_total": rice_total,
        "advance_received": advance_received,
        "rice_due": rice_due,

        "transport_rate_per_ton": transport_rate_per_ton,
        "transport_charge": transport_charge,
        "transport_paid_by_dealer": transport_paid_by_dealer,
        "transport_paid_by_customer": transport_paid_by_customer,
        "transport_due": transport_due,

        "breakup_rows": breakup_rows,
        "buy_cost_total": buy_cost_total,
        "profit_estimate": profit_estimate,
        "grand_total": grand_total,
    })


def sale_review(request):
    if request.method != "POST":
        return redirect("add_sale")

    # Get all form data
    data = request.POST.copy()

    bag_weight = Decimal(data.get("bag_weight") or "0")
    total_bags = Decimal(data.get("total_bags") or "0")
    rate_per_kg = Decimal(data.get("rate_per_kg") or "0")
    gst_percent = Decimal(data.get("gst_percent") or "0")
    advance = Decimal(data.get("advance_received") or "0")

    total_kg = bag_weight * total_bags
    taxable = total_kg * rate_per_kg
    gst_amt = (taxable * gst_percent) / Decimal("100")
    rice_total = taxable + gst_amt
    rice_due = rice_total - advance

    total_ton = total_kg / Decimal("1000")
    transport_rate = Decimal(data.get("transport_rate_per_ton") or "0")
    transport_charge = total_ton * transport_rate

    transport_due = transport_charge - (
        Decimal(data.get("transport_paid_by_dealer") or "0") +
        Decimal(data.get("transport_paid_by_customer") or "0")
    )

    context = {
        "data": data,
        "total_kg": round(total_kg, 2),
        "taxable": round(taxable, 2),
        "gst_amt": round(gst_amt, 2),
        "rice_total": round(rice_total, 2),
        "rice_due": round(rice_due, 2),
        "transport_charge": round(transport_charge, 2),
        "transport_due": round(transport_due, 2),
    }

    return render(request, "core/sale_review.html", context)

@transaction.atomic
def sale_confirm_save(request):
    if request.method != "POST":
        return redirect("sale_list")

    invoice_no = generate_sale_invoice_no()

    sale = Sale.objects.create(
        invoice_no=invoice_no,
        customer_name=request.POST.get("customer_name"),
        customer_gst=request.POST.get("customer_gst"),
        broker_id=request.POST.get("broker_id") or None,
        sale_date=request.POST.get("sale_date"),

        vehicle_number=request.POST.get("vehicle_number"),
        driver_name=request.POST.get("driver_name"),
        transporter_name=request.POST.get("transporter_name"),

        total_quantity_kg=request.POST.get("total_kg"),
        taxable_amount=request.POST.get("taxable"),
        gst_percent=request.POST.get("gst_percent"),
        gst_amount=request.POST.get("gst_amt"),

        transport_charge=request.POST.get("transport_charge"),
        transport_paid_by_dealer=request.POST.get("transport_paid_by_dealer") or 0,
        transport_paid_by_customer=request.POST.get("transport_paid_by_customer") or 0,

        advance_received=request.POST.get("advance_received") or 0,
        total_amount=request.POST.get("rice_total"),
        balance_amount=request.POST.get("rice_due"),
    )

    return redirect("sale_print", sale_id=sale.id)

def sale_print(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)

    rice_total = float(sale.taxable_amount) + float(sale.gst_amount)

    return render(request, "core/sale_print.html", {
        "sale": sale,
        "rice_total": rice_total,
    })



def broker_list(request):
    q = request.GET.get("q", "").strip()

    brokers = Broker.objects.all().order_by("-created_at")
    if q:
        brokers = brokers.filter(broker_name__icontains=q)

    return render(request, "core/broker_list.html", {
        "brokers": brokers
    })


def add_broker(request):
    if request.method == "POST":
        Broker.objects.create(
            broker_name=request.POST.get("broker_name"),
            mobile=request.POST.get("mobile", ""),
            gst_number=request.POST.get("gst_number", ""),
            opening_balance=request.POST.get("opening_balance") or 0,
            address=request.POST.get("address", "")
        )

        messages.success(request, "✅ Broker saved successfully!")
        return redirect("broker_list")

    return render(request, "core/add_broker.html")


def broker_report_detail(request, broker_id):
    broker = get_object_or_404(Broker, id=broker_id)

    sales = Sale.objects.filter(broker=broker).order_by("-sale_date", "-id")

    total_sales = sales.aggregate(s=Sum("total_amount"))["s"] or 0
    total_advance = sales.aggregate(s=Sum("advance_received"))["s"] or 0
    total_due = float(total_sales) - float(total_advance)

    return render(request, "core/broker_report_detail.html", {
        "broker": broker,
        "sales": sales,
        "total_sales": total_sales,
        "total_advance": total_advance,
        "total_due": total_due
    })


from io import BytesIO
from decimal import Decimal
import qrcode

from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth

from .models import Sale, SaleItem



def _amount_words(n: Decimal):
    number = int(n)
    words = num2words(number, lang='en_IN')
    return words.title() + " Only"

